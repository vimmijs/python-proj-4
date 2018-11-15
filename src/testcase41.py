from selenium import webdriver
import openpyxl

#load workbook

wk = openpyxl.load_workbook('C:\\Vimmi\\Job\\Study\\Python\\files\\testdata.xlsx')

print(wk.sheetnames)
print("Active sheet is " + wk.active.title)

sh = wk['MySheet1']
print(sh.title)
print(sh['A3'].value)

oCell = sh.cell(1,2)
print(oCell.value)

oCell = sh.cell(column=2, row=1)
print(oCell.value)
print(oCell.row)

rows = sh.max_row
cols = sh.max_column

print('total row ' + str(rows))
print('total cols ' + str(cols))

for i in range(1, rows+1):
    for j in range(1, cols+1):
        c = sh.cell (i,j)
        print(c.value)

print('new style')
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)
